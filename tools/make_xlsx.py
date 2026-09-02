#!/usr/bin/env python3
"""Build the formatted Excel workbook from the extracted gate-valve point list.

    python3 tools/make_xlsx.py

Reads data/gate_valve_coordinates.csv (produced by tools/gv_extract.py) and
writes data/gate_valve_coordinates.xlsx: the coordinate list on one sheet, and
the source, assumptions and points-to-check on another.

The count cells on the Notes sheet are formulas over the data range rather than
numbers typed in, so they stay right if rows are added or filtered. The
workbook is flagged to recalculate on open, which is what populates them --
openpyxl writes formulas without cached values.
"""


import csv

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROWS = list(csv.DictReader(open('data/gate_valve_coordinates.csv')))
SHEET = 'Gate Valve Coordinates'

FONT = 'Arial'
NAVY   = '1F3864'
HEADBG = '1F3864'
BANDBG = 'F2F5FA'
FLAGBG = 'FFF2CC'
thin = Side(style='thin', color='B4C6E7')
box  = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
ws = wb.active
ws.title = SHEET

# --- title block ---
ws.merge_cells('A1:D1')
ws['A1'] = 'GATE VALVE COORDINATES'
ws['A1'].font = Font(name=FONT, size=14, bold=True, color=NAVY)
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 24

ws.merge_cells('A2:D2')
ws['A2'] = '25_114D_NOC_WAT_13.0_RevF.dwg  ·  extracted from the gate-valve blocks in model space'
ws['A2'].font = Font(name=FONT, size=9, italic=True, color='595959')
ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 16

# --- header ---
HEAD = ['POINTS', 'EASTING', 'NORTHING', 'REMARKS']
for c, h in enumerate(HEAD, start=1):
    cell = ws.cell(row=4, column=c, value=h)
    cell.font = Font(name=FONT, size=10, bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor=HEADBG)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = box
ws.row_dimensions[4].height = 20

# --- data ---
first = 5
for i, r in enumerate(ROWS):
    row = first + i
    flagged = bool(r['REMARKS'].strip())
    band = FLAGBG if flagged else (BANDBG if i % 2 else 'FFFFFF')

    a = ws.cell(row=row, column=1, value=r['POINTS'])
    a.font = Font(name=FONT, size=10, bold=True)
    a.alignment = Alignment(horizontal='left')

    for col, key in ((2, 'EASTING'), (3, 'NORTHING')):
        cell = ws.cell(row=row, column=col, value=float(r[key]))
        cell.number_format = '0.000'
        cell.font = Font(name=FONT, size=10)
        cell.alignment = Alignment(horizontal='right')

    d = ws.cell(row=row, column=4, value=r['REMARKS'] or None)
    d.font = Font(name=FONT, size=9, italic=True, color='7F6000' if flagged else '000000')

    for col in range(1, 5):
        cell = ws.cell(row=row, column=col)
        cell.border = box
        cell.fill = PatternFill('solid', fgColor=band)

last = first + len(ROWS) - 1

ws.freeze_panes = 'A5'
ws.auto_filter.ref = f'A4:D{last}'
for col, width in (('A', 12), ('B', 15), ('C', 16), ('D', 42)):
    ws.column_dimensions[col].width = width
ws.print_title_rows = '4:4'

# --- notes sheet, counts driven by formulas over the data ---
nb = wb.create_sheet('Notes')
nb['A1'] = 'GATE VALVE COORDINATE LIST — SOURCE AND ASSUMPTIONS'
nb['A1'].font = Font(name=FONT, size=12, bold=True, color=NAVY)

rows = [
    ('Source drawing',      '25_114D_NOC_WAT_13.0_RevF.dwg'),
    ('Extracted from',      'Inserts of block CI_PW_GVN_PROP on layer PW_GV, model space'),
    ('Point names',         'The GV<n> text label nearest each valve, matched on the drawing’s label offset'),
    ('Coordinates',         'Each valve block’s insertion point, in the survey grid the drawing is modelled on'),
    ('Units',               'Metres, as held in the drawing. No transform applied — this drawing is modelled on the grid'),
    ('Extraction tool',     'tools/gv_extract.py in this repository'),
]
r = 3
for label, value in rows:
    nb.cell(row=r, column=1, value=label).font = Font(name=FONT, size=10, bold=True)
    nb.cell(row=r, column=2, value=value).font = Font(name=FONT, size=10)
    nb.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical='top')
    r += 1

r += 1
nb.cell(row=r, column=1, value='COUNTS').font = Font(name=FONT, size=11, bold=True, color=NAVY)
r += 1
count_row = {}
counts = [
    ('total', 'Points listed',            f"=COUNTA('{SHEET}'!A{first}:A{last})"),
    ('auto',  'Auto-numbered (no label)', f"=COUNTIF('{SHEET}'!D{first}:D{last},\"auto-numbered*\")"),
]
for key, label, formula in counts:
    nb.cell(row=r, column=1, value=label).font = Font(name=FONT, size=10, bold=True)
    c = nb.cell(row=r, column=2, value=formula)
    c.font = Font(name=FONT, size=10)
    c.number_format = '0'
    count_row[key] = r
    r += 1
nb.cell(row=r, column=1, value='Carrying a GV label').font = Font(name=FONT, size=10, bold=True)
c = nb.cell(row=r, column=2, value=f"=B{count_row['total']}-B{count_row['auto']}")
c.font = Font(name=FONT, size=10)
c.number_format = '0'
r += 1

r += 1
nb.cell(row=r, column=1, value='POINTS TO CHECK').font = Font(name=FONT, size=11, bold=True, color=NAVY)
r += 1
notes = [
    'GV174 is not a drawing label. The drawing holds 174 gate-valve blocks but only 173 GV labels; '
    'this valve has none and is absent from the drawing’s own table. It was given the next free '
    'number so it would not be dropped. Confirm whether it is a scheduled point before issuing.',

    'GV5 is missing from the drawing’s own table, where two consecutive rows are both labelled GV6. '
    'The drawing itself is right — only that table row is wrong. This list uses the drawing.',

    'Six points differ from the drawing’s table by 37–118 mm (GV135, GV87, GV67, GV62, GV116, GV83). '
    'This list follows the block positions. Whether that matters depends on your setting-out tolerance.',
]
for n in notes:
    nb.cell(row=r, column=1, value='•').font = Font(name=FONT, size=10)
    c = nb.cell(row=r, column=2, value=n)
    c.font = Font(name=FONT, size=10)
    c.alignment = Alignment(wrap_text=True, vertical='top')
    nb.row_dimensions[r].height = 42
    r += 1

nb.column_dimensions['A'].width = 24
nb.column_dimensions['B'].width = 96

wb.calculation.fullCalcOnLoad = True

out = 'data/gate_valve_coordinates.xlsx'
wb.save(out)
print('wrote', out, '|', len(ROWS), 'points, rows', first, '-', last)
