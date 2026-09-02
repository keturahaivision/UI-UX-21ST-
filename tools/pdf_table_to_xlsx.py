#!/usr/bin/env python3
"""Turn a plotted GATE VALVE COORDINATES table (PDF) into an Excel workbook.

    pdftotext -layout table.pdf table.txt
    python3 tools/pdf_table_to_xlsx.py table.txt -o gate_valve_callouts.xlsx

The PDF prints the table in two side-by-side blocks, so rows are matched
anywhere on a line rather than by column position.

Each row gets ready-made callout text in the drawing's own format --
`N=<northing>, E=<easting>` on one line, and the two halves separately so they
can be dropped straight into the N= / E= attributes of a callout block.

Pass --check with the CSV extracted from the drawing (tools/gv_extract.py) and
each row is compared against where its valve actually sits.
"""

from __future__ import annotations

import argparse
import csv
import math
import re
import sys

ROW_RE = re.compile(r"(GV\s?\d+)\s+(\d+\.\d+)\s+(\d+\.\d+)")


def read_rows(path: str):
    """Every GV row in the text dump, in printed order."""
    rows = []
    with open(path, encoding="utf-8", errors="replace") as fh:
        for line in fh:
            for m in ROW_RE.finditer(line):
                rows.append((m.group(1).replace(" ", ""),
                             float(m.group(2)), float(m.group(3))))
    return rows


def load_check(path: str):
    with open(path, newline="", encoding="utf-8") as fh:
        return {r["POINTS"]: (float(r["EASTING"]), float(r["NORTHING"]))
                for r in csv.DictReader(fh)}


def build(rows, check, out_path, precision=3):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    FONT = "Arial"
    NAVY, BAND, FLAG = "1F3864", "F2F5FA", "FFF2CC"
    thin = Side(style="thin", color="B4C6E7")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    SHEET = "Callouts"
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET

    ws.merge_cells("A1:G1")
    ws["A1"] = "GATE VALVE COORDINATES — CALLOUT SCHEDULE"
    ws["A1"].font = Font(name=FONT, size=14, bold=True, color=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:G2")
    ws["A2"] = ("Transcribed from the plotted coordinate table. "
                "CALLOUT is the text to place on the drawing.")
    ws["A2"].font = Font(name=FONT, size=9, italic=True, color="595959")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    head = ["POINTS", "EASTING", "NORTHING", "CALLOUT",
            "N LINE", "E LINE", "CHECKED AGAINST DRAWING"]
    for c, h in enumerate(head, start=1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = box
    ws.row_dimensions[4].height = 20

    # printed order is column-interleaved; present it in point order instead,
    # keeping both rows of a duplicated label
    rows = sorted(rows, key=lambda r: (int(r[0][2:]), r[1]))

    seen: dict[str, int] = {}
    first = 5
    for i, (name, e, n) in enumerate(rows):
        row = first + i
        seen[name] = seen.get(name, 0) + 1

        n_line = f"N={n:.{precision}f}"
        e_line = f"E={e:.{precision}f}"
        remark = ""

        if check:
            hit = check.get(name)
            if seen[name] > 1 or name not in check:
                remark = "duplicate label — see Notes"
            elif hit:
                d = math.hypot(hit[0] - e, hit[1] - n)
                remark = "matches" if d <= 0.02 else f"differs by {d*1000:.0f} mm"
            else:
                remark = "no matching point"

        flagged = bool(remark) and remark != "matches"
        fill = FLAG if flagged else (BAND if i % 2 else "FFFFFF")

        vals = [name, e, n, f"{n_line}, {e_line}", n_line, e_line, remark]
        for col, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = box
            cell.fill = PatternFill("solid", fgColor=fill)
            if col in (2, 3):
                cell.number_format = f"0.{'0'*precision}"
                cell.font = Font(name=FONT, size=10)
                cell.alignment = Alignment(horizontal="right")
            elif col in (4, 5, 6):
                cell.font = Font(name=FONT, size=10)
            elif col == 1:
                cell.font = Font(name=FONT, size=10, bold=True)
            else:
                cell.font = Font(name=FONT, size=9, italic=True,
                                 color="7F6000" if flagged else "595959")

    last = first + len(rows) - 1
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:G{last}"
    for col, width in (("A", 11), ("B", 14), ("C", 15), ("D", 34),
                       ("E", 18), ("F", 17), ("G", 26)):
        ws.column_dimensions[col].width = width
    ws.print_title_rows = "4:4"

    # ---- notes ----
    nb = wb.create_sheet("Notes")
    nb["A1"] = "HOW TO USE THIS SHEET"
    nb["A1"].font = Font(name=FONT, size=12, bold=True, color=NAVY)

    r = 3
    for label, value in [
        ("CALLOUT", "The whole callout on one line, in the drawing's format."),
        ("N LINE / E LINE", "The same text split in two, matching the two lines of a "
                            "callout box — paste straight into the N= and E= attributes."),
        ("CHECKED AGAINST DRAWING",
         "Each row compared with where its valve actually sits in the model."),
        ("Source", "Coordinates transcribed from the plotted table, not re-surveyed."),
    ]:
        nb.cell(row=r, column=1, value=label).font = Font(name=FONT, size=10, bold=True)
        c = nb.cell(row=r, column=2, value=value)
        c.font = Font(name=FONT, size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    r += 1
    nb.cell(row=r, column=1, value="COUNTS").font = Font(name=FONT, size=11, bold=True, color=NAVY)
    r += 1
    total_row = r
    for label, formula in [
        ("Rows in the table", f"=COUNTA('{SHEET}'!A{first}:A{last})"),
        ("Rows that match the drawing", f"=COUNTIF('{SHEET}'!G{first}:G{last},\"matches\")"),
    ]:
        nb.cell(row=r, column=1, value=label).font = Font(name=FONT, size=10, bold=True)
        c = nb.cell(row=r, column=2, value=formula)
        c.font = Font(name=FONT, size=10)
        c.number_format = "0"
        r += 1
    nb.cell(row=r, column=1, value="Rows needing attention").font = Font(name=FONT, size=10, bold=True)
    c = nb.cell(row=r, column=2, value=f"=B{total_row}-B{total_row+1}")
    c.font = Font(name=FONT, size=10)
    c.number_format = "0"
    r += 2

    nb.cell(row=r, column=1, value="BEFORE YOU USE THESE").font = Font(name=FONT, size=11, bold=True, color=NAVY)
    r += 1
    for note in [
        "The table has no GV5. Two rows are both labelled GV6, and the one at "
        "E 477853.859 / N 2744393.148 is really GV5 — it is the row flagged as a "
        "duplicate. The drawing itself carries a correct GV5 label; only the table row "
        "is wrong. GVFIXTABLE corrects it in the DWG.",

        "The table has no row for the 174th valve, at E 477149.577, N 2744103.507, which "
        "carries no GV label in the drawing either. Confirm whether it is a scheduled point.",

        "Rows marked 'differs by' sit 37–118 mm from their valve. The table was typed up "
        "before a later nudge of the symbol; the drawing is the newer of the two.",

        "Placing these by hand is what put a neighbour's coordinates in GV4's callout box "
        "in the drawing. GVANNO writes all of them straight from the block positions.",
    ]:
        nb.cell(row=r, column=1, value="•").font = Font(name=FONT, size=10)
        c = nb.cell(row=r, column=2, value=note)
        c.font = Font(name=FONT, size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        nb.row_dimensions[r].height = 42
        r += 1

    nb.column_dimensions["A"].width = 26
    nb.column_dimensions["B"].width = 96

    wb.calculation.fullCalcOnLoad = True
    wb.save(out_path)
    return len(rows)


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("text", help="pdftotext -layout output of the plotted table")
    ap.add_argument("-o", "--output", default="gate_valve_callouts.xlsx")
    ap.add_argument("--check", help="CSV from gv_extract.py, to compare each row against")
    ap.add_argument("--precision", type=int, default=3)
    args = ap.parse_args(argv)

    rows = read_rows(args.text)
    if not rows:
        print("no GV rows found — check the text dump", file=sys.stderr)
        return 1
    check = load_check(args.check) if args.check else None
    n = build(rows, check, args.output, args.precision)
    print(f"wrote {n} rows to {args.output}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
